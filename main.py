import datetime
from openpyxl import Workbook
import time


if __name__ == '__main__':
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = "Current Date"
    ws.cell(row=1, column=2).value = "Time"
    for i in range(2,5):
        ws.cell(row=i,column=1).value = datetime.datetime.now().strftime('%Y-%m-%d')
        ws.cell(row=i, column=2).value = datetime.datetime.now().strftime('%H:%M:%S')
        time.sleep(1)
    wb.save('gfg.xlsx')
    wb.close()